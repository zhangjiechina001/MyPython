import requests
import datetime
from openpyxl import load_workbook
import time
headers = {
            'User-Agent': 'User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.132 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.8',
            }
body = {
        'iActivityId':	'253767',
        'iFlowId' :	'603921',
        'g_tk':'1842395457'
        }
url = 'https://smoba.ams.game.qq.com/ams/ame/amesvr?ameVersion=0.3&sServiceType=yxzj&iActivityId=253767&sServiceDepartment=group_g&sSDID=2e8cf6296c34d41d3b0192633c0c922b&sMiloTag=AMS-MILO-253767-603921-unknown-1570786290653-zSDsLU&_=1570786290662'
while True:
    time.sleep(20)
    data = requests.post(url,headers=headers,data=body)
    vote=data.text
    id_name = {'v_1': '武陵仙君-诸葛亮', 'v_2': '辉光之辰-后羿', 'v_3': '蜜橘之夏-公孙离', 'v_4': '魔法小厨娘-安琪拉', 'v_5': '永曜之星-杨戬',
               'v_6': '游园惊梦-甄姬','v_7':'遇见飞天-杨玉环','v_8':'白虎志-百里玄策','v_9':'冰霜恋舞曲-干将莫邪','v_10':'大圣娶亲-孙悟空',
               'v_11':'奇迹圣诞-蔡文姬','v_12':'瑞麟志-花木兰','v_13':'青龙志-铠','v_14':'玄武志-苏烈','v_15':'一生所爱-露娜',
               'v_16':'朱雀志-百里守约','v_17':'逐梦之光-东皇太一','v_18':'逐梦之星-马可波罗','v_19':'逐梦之翼-哪吒','v_20':'云端筑梦师-庄周',}
    #str直接转dict
    votedict = eval(vote)
    nameid = votedict['modRet']['sOutValue2']
    votesum = votedict['modRet']['sOutValue3']
    namelist = nameid.split('|')
    votelist = votesum.split('|')
    for i in range(len(votelist)):
        votelist[i] = int(votelist[i])
    for i in range(len(namelist)):
        value = namelist[i]
        namelist[i] = id_name[value]
    finaldict = {}
    for i in range(len(namelist)):
        finaldict[namelist[i]] = votelist[i]
    a = sorted(finaldict.items(),reverse = True,key=lambda x:x[1])
    timenow = str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    print(timenow)
    wb = load_workbook('E://pvpvote.xlsx')
    #按顺序创建sheet
    wss = wb.create_sheet()
    votetime = []
    votetime.append(timenow)
    #append不能传str，所以写成list
    wss.append(votetime)
    for i  in a :
        print(i)
        wss.append(i)
    wb.save('E://pvpvote.xlsx')
